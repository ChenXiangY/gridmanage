from io import BytesIO
import os
from re import L
from PIL import Image as pilimg
import pymysql
import docx
from docx import Document
from docx.text.paragraph import Paragraph
from docx.image.image import Image
from docx.parts.image import ImagePart
from docx.oxml.shape import CT_Picture
import openpyxl
db = pymysql.connect(host='localhost',port=3306,user='root',password='ad5226236',database='GridManage')
# prepare a cursor object using cursor() method
cursor = db.cursor()

# execute SQL query using execute() method.
cursor.execute("SELECT VERSION()")

# Fetch a single row using fetchone() method.
data = cursor.fetchone()

print("Database version : %s " % data)


wb = openpyxl.load_workbook('result1.xlsx')
sheet=wb.worksheets[0]
counter = 1
while(True):
    if(sheet.cell(counter,1).value is None):
        break
    sql = """
        insert into managers(fatherid,ownid,name,phone,profileImg,gridname) values (%s,%s,%s,%s,%s,%s)
        """
    param=(
        sheet.cell(counter,1).value[0:8],
        sheet.cell(counter,1).value,
        sheet.cell(counter,3).value,
        sheet.cell(counter,4).value,
        'http://localhost:8080/ManageProfileImg/defaultProfileImg.jpg',
        sheet.cell(counter,2).value
    )
    cursor.execute(sql,param)
    db.commit()
    print(param)
    counter = counter+1
